##################################################
###the home of this script is: github.com/ezgranet/mmp-calc
##suggestions and contributions are eagerly welcomed
#### CREDITS: ezgranet & ING
#############################################################
########## THIS SOFTWARE LICENSED UNDER THE 'MIT LICENCE'
##### SEE END OF FILE FOR FULL DETAILS
######################################################





from openpyxl import Workbook, load_workbook
from pip._vendor.distlib.compat import raw_input

#selecting input


fileImport = raw_input("File Path (remember to include .xlsx)?")
formula = raw_input("Formula? (For D'Hondt, please type: DH    For St Lague, please type: SL )")
if formula in ["DH", "dh", " DH", "dh ", " dh"]:
    formula = 1
elif formula in ["SL", "sl", " SL", "sl ", " sl"]:
    formula = 2
g = raw_input("Please enter the number of Seats: ")

#here iterations refers to the number of quota and seat calcuations...e.g.
# entering "500" will return 250 columns of seats and 250 columns of quotas
#in order to compensate for user input, the loop iterations are doubled

workbook = load_workbook(filename=str(fileImport))
sheet = workbook.active
h = 0

#iterate to find how many parties
for x in range(2, 9999):
    a = sheet.cell(row=x, column=2).value
    if a is None:
        h = x
        break
h = h-1
print(h)
#declaring list size
intlseats = [1] * int(h)
intlvotes = [1] * int(h)
quotaX = [1] * (int(h) + 1)
#used to control the while loop
shouldStop = True

finalRow = int(h) + 1
rowWrite = 'f'
alphabetChecker = 0
alphabetTrack = 0

alphabet = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u',
            'v', 'w', 'x',
            'y', 'z']
q = 4
r = 3
letter = 0
columnVal = 3
otherColumn = 2
holdVotes = [1] * int(h)
#get initial list of seats

print(h)
for x in range(2, finalRow):
    a = sheet.cell(row=x, column=otherColumn).value
    intlseats[x - 2] = a
#get the initial number of votes
for d in range(2, finalRow):
    l = sheet.cell(row=d, column=columnVal).value
    holdVotes[d - 2] = float(l)

rack = 0

while (shouldStop):
    track = 0
    #track monitors what number loop we are on
    if rack == 0:
        #rack is used because I was too lazy to name my variables, ensures this loop only runs once
        for y in range(2, finalRow):
            b = sheet.cell(row=y, column=columnVal).value
            intlvotes[y - 2] = float(b)
        #calculates the first quota
        for b in range(2, finalRow):
            z = float(intlseats[track])
            p = float(intlvotes[track])
            newNum = p / ((formula * z) + 1)
            quotaX[b - 2] = newNum
            track = track + 1
            sheet[str(alphabet[q]) + str(b)] = quotaX[b - 2]
        for b in range(2, finalRow):
            sheet[str(alphabet[r]) + str(b)] = intlseats[b - 2]
    else:
        #read previous quota row and transfer into a new list
        for y in range(2, finalRow):
            b = sheet.cell(row=y, column=columnVal).value
            intlvotes[y - 2] = float(b)
        #identify which row will hold the highest number of votes
        IntmaxVal = intlvotes.index(max(intlvotes))
        #increase the number of seats in that row by 1
        intlseats[IntmaxVal] = intlseats[IntmaxVal] + 1
        # print(intlseats[IntmaxVal]). Use this for debugging purposes only.
        intlvotes[IntmaxVal] = holdVotes[IntmaxVal] / (intlseats[IntmaxVal] * formula + 1)
        # print(intlvotes[IntmaxVal]) Use this for debugging purposes only.
        #
        #
        for b in range(2, finalRow):
            sheet[str(alphabet[q]) + str(b)] = intlvotes[b - 2]
            #write the next row, for votes
        for b in range(2, finalRow):
            sheet[str(alphabet[r]) + str(b)] = intlseats[b - 2]
            #write the next row, for seats
    #the below function is used in the event that there are more than 26 rows.
    if q > 0:
        for i in range(0, 2):
            alphabet.append(str(alphabet[letter]) + alphabet[alphabetChecker])
            alphabetChecker = alphabetChecker + 1
            #keeps track of what letter row we are one
            if alphabetChecker == 26:
                letter = letter + 1
                alphabetChecker = 0
                #resets count in the event that you go more than 26
            alphabetTrack = 0
    if q > ((int(g) * 2)+2):
        shouldStop = False
        #breaks the loop after the correct number of iterations


    q = q + 2
    #track number row for quotas
    r = r + 2
    #track number row for votes
    rack = rack + 1
    # print(columnVal)
    columnVal = columnVal + 2
    # print(rack)

#EXPORT FUNCTIONS
#BELOW

fileExport = raw_input("File Save Name?")
fileSaveType = raw_input(
    "What type of file would you like to save as (CSV, XLSX. Do not include a period in the extension name.)?")
fileExportTwo = raw_input("Save file via System Dialogue (optimized for mac), or file path? \n Type 'FILE' or 'SYS' ")
fileExportTwo = raw_input("File Path?")
workbook.save(fileExportTwo + fileExport + "." + fileSaveType)


#######################
####LICENCE:
#########################
########Copyright (c) 2020 ezgranet & ING
#Permission is hereby granted, free of charge, to any person obtaining a copy
#of this software and associated documentation files (the "Software"), to deal
#in the Software without restriction, including without limitation the rights
#to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
#copies of the Software, and to permit persons to whom the Software is
#furnished to do so, subject to the following conditions:

#The above copyright notice and this permission notice shall be included in all
#copies or substantial portions of the Software.

########THE SOFTWARE IS PROVIDED 'AS IS', WITHOUT WARRANTY OF ANY KIND, EXPRESS #OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
#FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
#AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
#LIABILITY, WHETHER IN AN ACTION OF CONTRACT,, TORT OR OTHERWISE, ARISING FROM,
#OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
#SOFTWARE.



