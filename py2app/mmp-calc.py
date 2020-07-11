import tkinter as tk
import random2
from time import sleep
from tkinter import W, filedialog, INSERT, END
import tkinter.font as tkFont
import openpyxl as openpyxl
from self import self
import easygui
import csv
from openpyxl import Workbook
import os

global fileUsing
global numSeats
global formula
global setName


def runProgram():
    file_name = self.fileUsing
    print(os.path.splitext(file_name))
    file_name = os.path.splitext(file_name)[1]
    print(file_name)
    if file_name == ".csv":
        csvfilename = self.fileUsing
        excelfilename = "Hello.xlsx"
        wb = Workbook()
        ws = wb.active
        easygui.msgbox("Please select the output for the temporary XLSX file. This is used so the program can access "
                       "and run. You do not need excel. The file can be deleted immediately following running the "
                       "program should you so choose.")
        filenameOnE = easygui.diropenbox()
        with open(csvfilename) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            for row in csv_reader:
                ws.append(row)
        wb.save(str(filenameOnE))
        workbook = openpyxl.load_workbook(os.path.abspath(str(filenameOnE) + "/" + excelfilename + "." + "xlsx"))
        # print(os.path.abspath("/Users/Desktop/"+excelfilename))
        fileImport = os.path.abspath(excelfilename)
    else:
        fileImport = self.fileUsing
        workbook = openpyxl.load_workbook(filename=str(fileImport))
    # fileImport = easygui.fileopenbox()
    # elif i in ["file", "FILE", " FILE", "FILE ", " file", "file "]:
    #     fileImport = raw_input("File Path?")

    g = int(self.numSeats)
    # formula = easygui.enterbox("Please enter what formula to use (DH or SL): ")
    formula = self.formula
    if formula in ["DH", "dh", " DH", "dh ", " dh"]:
        formula = 1
    elif formula in ["SL", "sl", " SL", "sl ", " sl"]:
        formula = 2
    print(fileImport)
    formula = 1

    sheet = workbook.active
    h = 0

    # iterate to find how many parties
    for x in range(2, 9999):
        a = sheet.cell(row=x, column=2).value
        if a is None:
            h = x
            break
    h = h - 1
    print(h)
    # declaring list size
    intlseats = [1] * int(h)
    intlvotes = [1] * int(h)
    quotaX = [1] * (int(h) + 1)
    # used to control the while loop
    shouldStop = True

    finalRow = int(h) + 1
    rowWrite = 'f'
    alphabetChecker = 0
    alphabetTrack = 0

    alphabet = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u',
                'v', 'w', 'x',
                'y', 'z']
    q = 4
    r = 3
    letter = 0
    columnVal = 3
    otherColumn = 2
    holdVotes = [1] * int(h)
    # get initial list of seats

    print(h)
    for x in range(2, finalRow):
        a = sheet.cell(row=x, column=otherColumn).value
        intlseats[x - 2] = a
    # get the initial number of votes
    for d in range(2, finalRow):
        l = sheet.cell(row=d, column=columnVal).value
        holdVotes[d - 2] = float(l)

    rack = 0

    while (shouldStop):
        track = 0
        # track monitors what number loop we are on
        if rack == 0:
            # rack is used because I was too lazy to name my variables, ensures this loop only runs once
            for y in range(2, finalRow):
                b = sheet.cell(row=y, column=columnVal).value
                intlvotes[y - 2] = float(b)
            # calculates the first quota
            for b in range(2, finalRow):
                z = float(intlseats[track])
                p = float(intlvotes[track])
                newNum = p / ((formula * z) + 1)
                quotaX[b - 2] = newNum
                track = track + 1
                sheet[str(alphabet[q]) + str(b)] = quotaX[b - 2]
            for b in range(2, finalRow):
                sheet[str(alphabet[r]) + str(b)] = intlseats[b - 2]
        else:
            # read previous quota row and transfer into a new list
            for y in range(2, finalRow):
                b = sheet.cell(row=y, column=columnVal).value
                intlvotes[y - 2] = float(b)
            # identify which row will hold the highest number of votes
            IntmaxVal = intlvotes.index(max(intlvotes))
            # increase the number of seats in that row by 1
            intlseats[IntmaxVal] = int(intlseats[IntmaxVal]) + 1
            # print(intlseats[IntmaxVal]). Use this for debugging purposes only.
            intlvotes[IntmaxVal] = holdVotes[IntmaxVal] / (intlseats[IntmaxVal] * formula + 1)
            # print(intlvotes[IntmaxVal]) Use this for debugging purposes only.
            #
            #
            for b in range(2, finalRow):
                sheet[str(alphabet[q]) + str(b)] = intlvotes[b - 2]
                # write the next row, for votes
            for b in range(2, finalRow):
                sheet[str(alphabet[r]) + str(b)] = intlseats[b - 2]
                # write the next row, for seats
        # the below function is used in the event that there are more than 26 rows.
        if q > 0:
            for i in range(0, 2):
                alphabet.append(str(alphabet[letter]) + alphabet[alphabetChecker])
                alphabetChecker = alphabetChecker + 1
                # keeps track of what letter row we are one
                if alphabetChecker == 26:
                    letter = letter + 1
                    alphabetChecker = 0
                    # resets count in the event that you go more than 26
                alphabetTrack = 0
        if q > ((int(g) * 2) + 4):
            shouldStop = False
            # breaks the loop after the correct number of iterations

        q = q + 2
        # track number row for quotas
        r = r + 2
        # track number row for votes
        rack = rack + 1
        # print(columnVal)
        columnVal = columnVal + 2
        # print(rack)

    # EXPORT FUNCTIONS
    # BELOW
    # THIS WHOLE NEXT SECTION IS FOR SHOW. PEOPLE RAISE EYEBROWS IF YOUR CODE GOES TOO FAST
    listOne = ["Rigging election", "Calculating Seats", "Tallying ballots", "Running subprocess alpha",
               "Installing chinese spyware", "Calculating Formula", "Contacting Illuminati Server",
               "Enjoying a nice glass of wine", "Billing for a few more hours of work", "Writing to spreadsheet",
               "Performing merge sort", "Taking a quick break",
               "Preparing", "Outputting.", "Outputting..", "Outputting...", "Outputting", "Outputting.", "Outputting..",
               "Outputting...", "Outputting.", "Done!"]
    textOne = tk.Text(root)
    textOne.pack()
    textOne.place(relx=.65, rely=.4)
    textOne.config(width=40, height=2)

    # EXPORT OPTIONS
    fileExport = self.setName
    fileSaveType = easygui.enterbox(
        "What type of file would you like to save as (CSV, XLSX. Do not include a period in the extension name.)?")
    # fileExportTwo = raw_input("Save file via System Dialogue (optimized for mac), or file path? \n Type 'FILE' or 'SYS' ")
    # if i in ["sys", "SYS", " SYS", "SYS ", " sys" "sys "]:
    fileExportTwo = easygui.diropenbox()
    # PLAYING THROUGH THE ABOVE MENTIONED MEANINGLESS LIST
    for x in listOne:
        textOne.insert(END, x)
        root.update()
        sleep((random2.random() * .8) + .05)
        print(textOne)
        if x == len(listOne) - 1:
            sleep(10)
        textOne.delete('1.0', END)

    workbook.save(str(fileExportTwo) + "/" + fileExport + "." + fileSaveType)

    # Developed by ING 2020. All rights reserved.
    # This software is intended to be open source and may be used or modified by anyone.


def UploadAction(event=None):
    filename = filedialog.askopenfilename()
    self.fileUsing = filename
    print('Selected:', filename)


def setSeats():
    self.numSeats = e2.get()
    print(self.numSeats)


def setFormula():
    self.formula = e3.get()
    print(self.formula)


def setName():
    self.setName = e4.get()
    print(self.setName)


root = tk.Tk()
root.title("mmp-calc")
root.geometry("800x600")
text = tk.Text(root)
root.update()
text.insert(INSERT, "loading...")
text.pack()
text.place(x=260, y=200)
text.config(fg="white")
fontStyle = tkFont.Font(family="Courier", size=20)
text.config(bg="blue", font=fontStyle, borderwidth=2, relief="groove")
text.tag_configure("center", justify='center')
text.tag_add("center", "1.0", "end")
text.config(width=20, height=2)

root.update()
sleep(3)
text.destroy()
fontStyleOne = tkFont.Font(family="Courier", size=30)
labelTop = tk.Label(root, text="mmp-calc")
labelTop.config(font=fontStyleOne)
labelTop.config(bg="blue", font=fontStyle, borderwidth=2, relief="groove")
labelTop.config(fg="white")
labelTop.pack()
runProgram = tk.Button(root, text='Run', command=runProgram, height=1, width=5)

uploadbutton = tk.Button(root, text='Open', command=UploadAction)
uploadbutton.pack()
uk = tk.Label(root, text="Open File")

uk.place(relx=0.1, rely=0.2, anchor=W)
uploadbutton.place(relx=0.2, rely=0.2, anchor=W)

ik = tk.Label(root, text="NumSeats")
ik.place(relx=0.1, rely=0.3, anchor=W)
e2 = tk.Entry(root)
e2.pack()
enterButton = tk.Button(root, text='enter', command=setSeats)
enterButton.pack()
enterButton.place(relx=0.4, rely=0.35, anchor=W)
e2.place(relx=0.1, rely=0.35, anchor=W)

jk = tk.Label(root, text="Formula? DH or SL")
jk.place(relx=0.1, rely=0.45, anchor=W)
e3 = tk.Entry(root)
e3.pack()
enterButtonTwo = tk.Button(root, text='enter', command=setFormula)
enterButtonTwo.pack()
enterButtonTwo.place(relx=0.4, rely=0.5, anchor=W)
e3.place(relx=0.1, rely=0.5, anchor=W)

pk = tk.Label(root, text="Save name?")
pk.place(relx=0.1, rely=0.55, anchor=W)
e4 = tk.Entry(root)
e4.pack()
enterButtonThree = tk.Button(root, text='enter', command=setName)
enterButtonThree.pack()
enterButtonThree.place(relx=0.4, rely=0.6, anchor=W)
e4.place(relx=0.1, rely=0.6, anchor=W)

runProgram.pack()
lk = tk.Label(root, text="Run Program")
lk.place(relx=0.1, rely=0.75, anchor=W)
runProgram.place(relx=0.3, rely=0.75, anchor=W)
runProgram.config(borderwidth=2)

tk.mainloop()
